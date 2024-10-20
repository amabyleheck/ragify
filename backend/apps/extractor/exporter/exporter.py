import os
import re

from typing import List, Dict, Union

from openpyxl import load_workbook  # Library for importing excel files

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from openpyxl.styles import Font

from langchain_core.documents import Document
from extractor.utils.utils import get_variable_form

ABS_PATH: str = os.path.dirname(os.path.abspath(__file__))
EXCEL_TEMPLATE_PATH = f"{ABS_PATH}/export_template.xlsx"


class ExcelResultsExport:
    """
    Classe principal de geração de resultados e de documento de saída. Recebe os dicionários de n-modelos
    e compara-os a partir do dicionário anotado importado.

    Atualmente, conta com três folhas:

    1. Resultados: Lista as porcentagens de acurácia de extração de modelo, onde cada linha lista a performance de determinado prompt:
        -

    |                   MODELO                    |
    |                 Llama-2-7b                  |
    |   ID PROMPT   |   VARIÁVEL    |    PROMPT   |


    2. Detalhes: Lista para cada documento o valor esperado, o valor extraído, e o resultado da comparação para número de edital e município de irregularidade.
    Utilizado para checagem de comparação e métricas (debugging).
    Caso a execução seja com mais de um modelo, a mesma folha contará com n-tabelas com detalhes de cada uma.

    |                            DETALHES DA COMPARAÇÃO                                 |
    |                |                     Nome da coluna escolhido                     |
    |  ID Documento  | Valor Esperado       | Valor Extraído         | Resultado        |


    3. Especificações: Tabela que lista todos os parâmetros setados para determinada extração.

    |                                                              ESPECIFICAÇÕES DA COMPARAÇÃO                                                                                                          |
    |        Modelo       |                                                Embeddings                                               |      Retrieval       |
    |  Nome | Quantização | Tamanho de chunk | Tamanho de chunk overlap | Modelo BERT | Estratégia de splitting  | Banco de vetores |  Chain Type | Top k  |


    :param annotated_dict: Dicionário anotado vindo da importação de excel. Tem como chaves os IDs de documentos.
    :param extracted_dict: Dicionário anotado vindo da importação de excel. Tem como chaves os modelos que realizaram a extração.

    """

    INITIAL_RESULTS_ROW = 4
    INITIAL_DETAIL_ROW = 4
    INITIAL_SPECS_ROW = 4

    INITIAL_RESULTS_COL = 0

    def __init__(
        self,
        documents_ids: List,
        annotated_dict: Dict,
        extracted_dict: Dict,
        variables: List,
        model_name: str,
    ) -> None:
        """
        Salva variáveis em atributos e realiza a comparação de todos os modelos.
        """

        self.variables = variables

        self.documents_ids = documents_ids
        self.total_parsed_documents = len(self.documents_ids)

        self.model_name = model_name

        self.annotated_dict = annotated_dict
        self.extracted_dict = extracted_dict

        self.resultado_geral_dict = {}

        # REGION: Comparando resultados para cada modelo e persistindo no dicionário
        for variable in self.variables:
            key = variable.get("name")
            # if label := variable.get("label"):
            #     key += f" - {label}"
            self.resultado_geral_dict[key] = self.compare(variable)
        # ENDREGION

    def compare(self, variable: dict) -> Dict:
        """
        Rotina de comparação de resultados extraídos X valores anotados.

        Invoca dois métodos de pós-processamento de número de edital e município que farão a validação dos resultados
        e calcula ao final valores de acurácia por variável e média total.

        :param model_name: String com nome de modelo com mesmo valor de chave do dicionário extraído

        :return: Retorna dicionário com valores extraídos, anotados e seu resultado de comparação para cada documento
        somando a uma chave `metricas` que calcula as porcentagens finais a serem utilizadas na página "Resultados"
        """
        variable_name = variable.get("name")
        variable_form = get_variable_form(variable_name)

        current_extracted_dict = self.extracted_dict[variable_name]
        current_annotated_dict = self.annotated_dict[variable_name]

        total_successes = 0
        total_retriever_successes = 0

        variable_result = {}

        # REGION: Comparando valor extraído e anotado para cada documento para folha `Detalhes`
        for document in self.documents_ids:

            key = document + ".pdf"
            annotated_value = current_annotated_dict[key]
            extracted_value = current_extracted_dict[document]["value"]
            retrieved_docs = current_extracted_dict[document]["docs"]

            if "processolicitatório" in variable_form:
                processed_variable = self.process_n_processo_licitatorio(
                    document, annotated_value, extracted_value
                )
                has_retriever_found_answer = self.process_n_processo_licitatorio(
                    document, annotated_value, retrieved_docs
                )
                if processed_variable and not has_retriever_found_answer:
                    # Comumente, a sequência exata está envolvida por caracteres especiais (e.g '\n'),
                    # o que muitas vezes nao impede o modelo de acertar a resposta
                    print("Modelo acertou mesmo sem sequência exata nos documentos.")
                    has_retriever_found_answer = (
                        True  # Evitar acurácia de retriever menor que acurácia obtida
                    )
            elif "municípiodeirregularidade" in variable_form:
                processed_variable = self.process_municipio(
                    document, annotated_value, extracted_value
                )
                has_retriever_found_answer = self.process_municipio(
                    document, annotated_value, retrieved_docs
                )
            else:
                processed_variable = self.general_processing(
                    document, annotated_value, extracted_value
                )
                has_retriever_found_answer = self.general_processing(
                    document, annotated_value, retrieved_docs
                )

            if processed_variable:
                total_successes += 1

            if has_retriever_found_answer:
                total_retriever_successes += 1

            # Final result dict
            variable_result[document] = {
                "valor_anotado": annotated_value,
                "valor_extraido": extracted_value,
                "resultado": processed_variable,
                "resultado_retriever": has_retriever_found_answer,
            }
        # ENDREGION

        # REGION: Calculando porcentagens finais de folha `Resultados`
        success_rate = round(total_successes / self.total_parsed_documents, 2)
        retriever_success_rate = round(
            total_retriever_successes / self.total_parsed_documents, 2
        )

        variable_result["metricas"] = {
            "success_rate": success_rate,
            "retriever_success_rate": retriever_success_rate,
        }
        # ENDREGION

        return variable_result

    def general_processing(
        self, document_id, annotated_value: str, extracted_value: str
    ) -> bool:
        # IMPLEMENTAR AQUI SEU MÉTODO
        if annotated_value.lower() in extracted_value.lower():
            return True
        else:
            return False

    @staticmethod
    def process_n_processo_licitatorio(
        document_id, annotated_value: str, extracted_value: str | List[Document]
    ) -> bool:
        """
        Pós-processamento para validação do valor extraído de número do processo licitatório de determinado documento através de Regex.

        :param document_id: Número do documento para logging
        :param annotated_value: Valor anotado para número edital (valor esperado)
        :param extracted_value: Valor extraído de número de edital

        :return bool: Retorna True caso número de edital extraído for o mesmo que o anotado, False caso contrário.
        """

        # QUALQUER SEQUÊNCIA NUMÉRICA + / + QUALQUER SEQUÊNCIA NUMÉRICA
        pattern = r"\d+/\d+"

        annotated_match = re.findall(pattern, annotated_value)[
            0
        ]  # Evitar discrepância de valores importados
        if not isinstance(extracted_value, list):
            extracted_value = [extracted_value]

        for text in extracted_value:
            if isinstance(text, Document):
                # Armazenando em text apenas o conteúdo do chunk retornado pelo retriever
                text = text.page_content
            extracted_matches = re.findall(pattern, text)
            for find in extracted_matches:
                if annotated_match in find:
                    return True

        return False

    @staticmethod
    def process_municipio(
        document_id, annotated_value: str, extracted_value: str
    ) -> bool:
        """
        Pós-processamento para validação do valor extraído de município de irregularidade de determinado documento.

        :param document_id: Número do documento para logging
        :param annotated_value: Valor anotado para número edital (valor esperado)
        :param extracted_value: Valor extraído de número de edital

        :return bool: Retorna True caso município de irregularidade extraído for o mesmo que o anotado, False caso contrário.
        """
        if annotated_value.lower() in extracted_value.lower():
            return True
        return False

    def fill_cell(
        self, value: bool, success_title: str = None, error_title: str = None
    ) -> Union[str, Font]:
        """
        Retorna palavra dependendo do resultado de validação de célula.

        :value: Booleano com valor de retorno do pós-processamento de variável
        :return: 'ERRO' em fonte vermelha caso validação seja falsa, 'ACERTO' em fonte verde caso validação seja bem sucedida
        """
        if not value:
            return "ERRO" if not error_title else error_title, Font(
                color="ff0000", bold=True
            )
        else:
            return "ACERTO" if not success_title else success_title, Font(
                color="31c120", bold=True
            )

    def create_details_sheets(self, workbook: Workbook) -> None:
        """
        Copia a folha de detalhes e cola uma após a outra, dependendo da quantidade de extrações, renomeando-as de acordo com o nome dos modelos.

        :param workbook: Documento excel onde será realizada a cópia
        """

        folha_detalhes = workbook["Detalhes"]

        for index, variable_dict in enumerate(self.variables):
            nome_folha = f'Detalhes - Variable {variable_dict.get("name")} - {variable_dict.get("label")}'
            if index == 0:
                folha_detalhes.name = nome_folha
                folha_detalhes.title = nome_folha
            else:
                copia_folha = workbook.copy_worksheet(folha_detalhes)
                copia_folha.name = nome_folha
                copia_folha.title = nome_folha

                # Movendo cópia para antes da folha de Especificações
                workbook.move_sheet(copia_folha, offset=-1)

    def export_specs(
        self, worksheet: Worksheet, model_kwargs: Dict, embedding_kwargs: Dict
    ) -> None:
        """
        Preenche a tabela de especificações com parâmetros de modelo, prompt e embeddings na folha `Especificações`.


        |                                                              ESPECIFICAÇÕES DA COMPARAÇÃO                                                                                                          |
        |        Modelo       |                                                Embeddings                                               |      Retrieval       |
        |  Nome | Quantização | Tamanho de chunk | Tamanho de chunk overlap | Modelo BERT | Estratégia de splitting  | Banco de vetores |  Chain Type | Top k  |


        :param worksheet: Folha de excel a ser realizada preenchimento (neste caso, sempre será a folha `Especificações`)
        :param model_name: Número do documento para logging
        :param embedding_kwargs: Valor extraído de número de edital
        """

        for row in worksheet.iter_rows(
            min_row=self.INITIAL_SPECS_ROW, max_row=self.INITIAL_SPECS_ROW
        ):
            # MODEL REGION
            row[0].value = model_kwargs.get("model_name")  # SETTING MODEL'S NAME CELL
            row[1].value = model_kwargs.get(
                "quantization"
            )  # SETTING MODEL'S QUANTIZATION
            # ENDREGION

            # EMBEDDINGS REGION
            row[2].value = embedding_kwargs.get("chunk_size")  # SETTING CHUNK SIZE CELL
            row[3].value = embedding_kwargs.get(
                "chunk_overlap"
            )  # SETTING CHUNK OVERLAP SIZE CELL
            row[4].value = embedding_kwargs.get("bert_model")  # SETTING BERT MODEL CELL
            row[5].value = embedding_kwargs.get(
                "splitter"
            )  # SETTING SPLITTING STRATEGY CELL
            row[6].value = embedding_kwargs.get(
                "vector_db_class"
            )  # SETTING SPLITTING STRATEGY CELL
            # ENDREGION

            # RETRIEVAL CHAIN REGION
            row[7].value = model_kwargs.get("chain_type")  # SETTING CHAIN TYPE CELL
            row[8].value = model_kwargs.get("top_k")  # SETTING CHAIN'S TOP K
            # ENDREGION

    def export_results_details(self, workbook: Workbook):
        """
        Preenche a tabela de detalhamento de comparação de documentos por modelo.

        |                            DETALHES DA COMPARAÇÃO                                 |
        |                |                     Nome da coluna escolhido                     |
        |  ID Documento  | Valor Esperado       | Valor Extraído         | Resultado        |

        :param worksheet: Folha de excel a ser realizada preenchimento (neste caso, sempre será a folha `Detalhes`)
        """

        # Caso existam mais de um prompt, copia-se os cabeçalhos da tabela detalhes n-vezes,
        # onde cada cabeçalho armazenará o resultado de um prompt

        self.create_details_sheets(workbook=workbook)

        last_id = self.documents_ids[-1]

        for index, variable_dict in enumerate(self.variables):
            nome_folha = f'Detalhes - Variable {variable_dict.get("name")} - {variable_dict.get("label")}'
            worksheet = workbook[nome_folha]

            # Preenchendo nome da coluna com o mesmo valor do arquivo importado
            worksheet.cell(2, 2).value = variable_dict.get("name")

            current_model_dict = self.resultado_geral_dict[variable_dict.get("name")]

            for index, row in enumerate(
                worksheet.iter_rows(
                    min_row=self.INITIAL_DETAIL_ROW,
                    max_row=self.INITIAL_DETAIL_ROW + self.total_parsed_documents,
                )
            ):
                current_document_id = self.documents_ids[index]

                current_document_results = current_model_dict[current_document_id]

                # ID REGION
                row[0].value = current_document_id  # SETTING DOCUMENT ID
                # ENDREGION

                # NÚMERO DE EDITAL REGION
                row[1].value = current_document_results["valor_anotado"]
                row[2].value = current_document_results["valor_extraido"]
                row[3].value, row[3].font = self.fill_cell(
                    current_document_results["resultado"]
                )
                row[4].value, row[4].font = self.fill_cell(
                    current_document_results["resultado_retriever"],
                    success_title="SIM",
                    error_title="NÃO",
                )
                # ENDREGION

                if current_document_id == last_id:
                    break

    def export_results(self, worksheet: Worksheet):
        """
        Preenche a tabela de métricas de acurácia por modelo.

        |                   MODELO                    |
        |                 Llama-2-7b                  |
        |   ID PROMPT   |   VARIÁVEL    |    PROMPT   |

        :param worksheet: Folha de excel a ser realizada preenchimento (neste caso, sempre será a folha `Resultados`)
        """
        default_style = None
        # Preenchendo nome da coluna com o mesmo valor do arquivo importado
        worksheet.cell(2, 1).value = self.model_name

        for index, row in enumerate(
            worksheet.iter_rows(
                min_row=self.INITIAL_RESULTS_ROW,
                min_col=self.INITIAL_RESULTS_COL,
                max_row=self.INITIAL_RESULTS_ROW + len(list(self.variables)),
            )
        ):
            try:
                current_variable = self.variables[index]
                current_variable_name = current_variable.get("name")
                current_variable_prompt = current_variable.get("prompt")
            except IndexError:
                break

            current_result_dict = self.resultado_geral_dict[current_variable_name]
            if index == 0:
                default_style = row[0].style
            else:
                row[0].style = default_style
            row[0].value = current_variable_name
            row[1].value = current_result_dict["metricas"]["success_rate"]
            row[2].value = current_variable_prompt
            row[3].value = current_result_dict["metricas"]["retriever_success_rate"]

    def export(
        self, model_kwargs: Dict, embedding_kwargs: Dict, directory: str
    ) -> Dict:
        """
        Método a ser chamado após instanciação da classe. Invoca todas as funções de preenchimento das folhas.

        :param model_kwargs: Dicionário contendo informações de modelo (nome, quantização e top k)
        :param embedding_kwargs: Dicionário contendo informações de embeddings
        (tamanho de chunk, tamanho de chunk overlap, modelo BERT, splitter e banco de vetores)
        """
        workbook = load_workbook(os.path.abspath(EXCEL_TEMPLATE_PATH))

        folha_resultados = workbook.get_sheet_by_name("Resultado")
        folha_especificacoes = workbook.get_sheet_by_name("Especificações")

        # PÁGINA RESULTADOS REGION
        self.export_results(worksheet=folha_resultados)
        # ENDREGION

        # PÁGINA DETALHES REGION
        self.export_results_details(workbook=workbook)
        # ENDREGION

        # PÁGINA ESPECIFICAÇÕES REGION
        self.export_specs(
            worksheet=folha_especificacoes,
            model_kwargs=model_kwargs,
            embedding_kwargs=embedding_kwargs,
        )
        # ENDREGION

        # EXPORTANDO ARQUIVO FINAL REGION
        model_name = model_kwargs.get("model_name")
        chunk_size = embedding_kwargs.get("chunk_size")
        chunk_overlap = embedding_kwargs.get("chunk_overlap")
        top_k = model_kwargs.get("top_k")

        variables_section = f"chunk_s{chunk_size}-o{chunk_overlap}_top_k{top_k}"

        if directory:
            os.makedirs(directory, exist_ok=True)

        results_file_name = f"RESULTADOS_{model_name}__{variables_section}.xlsx"
        file_name = directory + results_file_name
        # ENDREGION

        workbook.save(filename=file_name)

        return self.get_results_dict_from_export(file_name)

    @staticmethod
    def get_results_dict_from_export(file_path: str) -> Dict:
        """Retorna um dicionário com todas as informações úteis de um resultado de uma extração (produzida pela mesma classe)

        Args:
            file_path (str): Caminho de arquivo .xslx provindo de uma execução de extração

        Returns:
            Dict: Retorna um dicionário com chaves representando os principais parâmetros e resultados de uma extração.
        """
        workbook = load_workbook(file_path)

        folha_resultados = workbook.get_sheet_by_name("Resultado")
        folha_especificacoes = workbook.get_sheet_by_name("Especificações")

        return {
            "model": folha_resultados.cell(2, 1).value,
            "accuracy": folha_resultados.cell(4, 2).value * 100,
            "retriever_accuracy": folha_resultados.cell(4, 4).value * 100,
            "chunk_size": folha_especificacoes.cell(4, 3).value,
            "chunk_overlap": folha_especificacoes.cell(4, 4).value,
            "top_k": folha_especificacoes.cell(4, 9).value,
            "bert_model": folha_especificacoes.cell(4, 5).value,
            "chain_type": folha_especificacoes.cell(4, 7).value,
        }
