from typing import List
from openpyxl import load_workbook  # Library for importing excel files


class EmptyCellException(Exception):
    """Exceção a ser levantada em caso de células vazias."""


class ExcelImportParser:
    file_path: str = "import.xlsx"
    """
    Carrega arquivo xlsx preenchido a partir de template com arquivos escolhidos e anotados
    Cada linha deve conter ID de arquivo e uma coluna com a variável escolhida para a extração

    Salva em self.ANNOTATED_DICT um dicionário onde cada chave é um ID de arquivo e suas respectivas informações
    """
    INITIAL_ROW = 2

    def __init__(self) -> None:

        workbook = load_workbook(self.file_path)
        main_sheet = workbook.worksheets[0]
        self.variable_name = self.get_column_header(worksheet=main_sheet)

        self.ANNOTATED_DICT = {}

        for row in main_sheet.iter_rows(min_row=self.INITIAL_ROW):
            id = self.parse_ID(row[0].value)
            valor = self.parse_str(row[1].value, column_name=self.variable_name)

            self.ANNOTATED_DICT[id] = {"value": valor}

    def get_annotated_dict(self):
        return self.ANNOTATED_DICT

    def get_column_name(self):
        return self.variable_name

    def get_ids_documentos(self) -> List[int]:
        return list(self.get_annotated_dict().keys())

    def get_column_header(self, worksheet) -> str:
        """Retorna o nome da variável a ser extraída, mesmo nome preenchido pelo usuário no arquivo importado

        Returns:
            str: String extraída do nome da coluna preenchida pelo usuário
        """
        return worksheet.cell(1, 2).value

    @staticmethod
    def get_column_error_message(column_name: str):
        return f"Coluna {column_name} é obrigatória."

    def parse_ID(self, data) -> int:
        """
        Processa colunas de ID, que devem ser sempre numéricas. Caso coluna esteja vazia ou não numérica, levanta uma exceção.
        """
        try:
            value = int(data)
            if not value:
                raise EmptyCellException(self.get_column_error_message("ID"))
        except ValueError:
            raise Exception("Valor inserido em coluna ID deve ser numérico.")

        return value

    def parse_str(self, data, column_name: str) -> int:
        """
        Processa restante das colunas, que devem ser sempre strings. Caso coluna esteja vazia, levanta uma exceção.
        """
        if not data:
            raise EmptyCellException(self.get_column_error_message(column_name))

        return data
